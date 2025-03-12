import tkinter as tk
from tkinter import ttk, messagebox, filedialog, colorchooser
import re
import os
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from database import init_db, get_user, add_user, init_pacientes_db, cargar_desde_base_de_datos
from utils import validar_rut, obtener_codigo_barras, cargar_listado_examenes_desde_excel, obtener_codigo_paciente
import appointments
import reports
import communication

# Variables globales
pacientes = {}
# Verificar si el archivo Excel existe y cargar los exámenes disponibles.
ruta_excel = os.path.join(os.path.dirname(__file__), "MAESTRO_DE_EXAMENES.xlsx")
if os.path.exists(ruta_excel):
    examenes_disponibles = cargar_listado_examenes_desde_excel(ruta_excel)
else:
    examenes_disponibles = []
    messagebox.showerror("Error", "El archivo MAESTRO_DE_EXAMENES.xlsx no se encontró en el directorio del proyecto.")

# ---------------------- VENTANA DE LOGIN ----------------------
def login_ui():
    login_win = tk.Tk()
    login_win.geometry("250x150")
    login_win.title("Usuarios")

    # Menú para opciones de color y comunicación
    menubar = tk.Menu(login_win)  # Cambio 'root' por 'login_win'
    login_win.config(menu=menubar)
    options_menu = tk.Menu(menubar, tearoff=False)
    options_menu.add_command(label="Cambiar Color Texto", command=lambda: change_text_color(login_win))
    options_menu.add_command(label="Cambiar Color Fondo", command=lambda: change_bg_color(login_win))
    # Agregar el comando para iniciar la comunicación con los dispositivos
    options_menu.add_command(label="Iniciar Comunicación", command=lambda: communication.start_device_communication())
    menubar.add_cascade(label="Opciones", menu=options_menu)

    tk.Label(login_win, text="Email:").grid(row=0, column=0, padx=10, pady=5)
    email_entry = tk.Entry(login_win)
    email_entry.grid(row=0, column=1, padx=10, pady=5)
    tk.Label(login_win, text="Contraseña:").grid(row=1, column=0, padx=10, pady=5)
    contraseña_entry = tk.Entry(login_win, show="*")
    contraseña_entry.grid(row=1, column=1, padx=10, pady=5)
    
    tk.Button(login_win, text="Registrarse", command=lambda: register_user_ui(login_win)).grid(row=2, column=0, columnspan=2, padx=10, pady=5)
    tk.Button(login_win, text="Iniciar sesión", command=lambda: login_user_action(login_win, email_entry, contraseña_entry)).grid(row=3, column=0, columnspan=2, padx=10, pady=5)
    
    login_win.mainloop()


def register_user_ui(parent):
    top = tk.Toplevel(parent)
    top.title("Crear cuenta")
    tk.Label(top, text="Nombre:").grid(row=0, column=0, padx=10, pady=5)
    nombre_entry = tk.Entry(top)
    nombre_entry.grid(row=0, column=1, padx=10, pady=5)
    tk.Label(top, text="Email:").grid(row=1, column=0, padx=10, pady=5)
    email_entry = tk.Entry(top)
    email_entry.grid(row=1, column=1, padx=10, pady=5)
    tk.Label(top, text="Contraseña:").grid(row=2, column=0, padx=10, pady=5)
    contraseña_entry = tk.Entry(top, show="*")
    contraseña_entry.grid(row=2, column=1, padx=10, pady=5)
    
    def close_and_register():
        if add_user(nombre_entry.get(), email_entry.get(), contraseña_entry.get()):
            messagebox.showinfo("Cuenta creada", "La cuenta ha sido creada correctamente")
            top.destroy()
        else:
            messagebox.showerror("Error", "El email ingresado ya existe")
    
    tk.Button(top, text="Crear cuenta", command=close_and_register).grid(row=3, column=0, columnspan=2, padx=10, pady=5)

def login_user_action(login_win, email_entry, contraseña_entry):
    user = get_user(email_entry.get(), contraseña_entry.get())
    if user:
        messagebox.showinfo("Inicio de sesión", f"Bienvenido {user[1]}")
        login_win.destroy()
        main_ui()
    else:
        messagebox.showerror("Error", "El email o la contraseña son incorrectos")

def change_text_color(win):
    color = colorchooser.askcolor(title="Elige un color para el texto")[1]
    if color:
        win.configure(fg=color)

def change_bg_color(win):
    color = colorchooser.askcolor(title="Elige un color para el fondo")[1]
    if color:
        win.configure(bg=color)

# ---------------------- VENTANA PRINCIPAL (SISTEMA) ----------------------
def main_ui():
    # Crear la ventana principal
    root = tk.Tk()  # Asegúrate de que 'root' esté definido aquí
    root.title("Sistema (R.P.E.D)")

    # Menú para opciones de color y comunicación
    menubar = tk.Menu(root)
    root.config(menu=menubar)
    options_menu = tk.Menu(menubar, tearoff=False)
    options_menu.add_command(label="Cambiar Color Texto", command=lambda: change_text_color(root))
    options_menu.add_command(label="Cambiar Color Fondo", command=lambda: change_bg_color(root))
    # Agregar el comando para iniciar la comunicación con los dispositivos
    options_menu.add_command(label="Iniciar Comunicación", command=lambda: communication.start_device_communication())
    menubar.add_cascade(label="Opciones", menu=options_menu)

    notebook = ttk.Notebook(root)
    notebook.pack(fill=tk.BOTH, expand=True)

    registro_pacientes = ttk.Frame(notebook)
    registro_examenes = ttk.Frame(notebook)
    listado_derivacion = ttk.Frame(notebook)
    acerca_de = ttk.Frame(notebook)

    notebook.add(registro_pacientes, text="Registro de Pacientes")
    notebook.add(registro_examenes, text="Registro de Exámenes")
    notebook.add(listado_derivacion, text="Listado de Derivación")
    notebook.add(acerca_de, text="Acerca de")

    # Variables del formulario
    global nombre_var, rut_var, fecha_nacimiento_var, edad_var, sexo_var, examen_var, resultado_var
    nombre_var = tk.StringVar()
    rut_var = tk.StringVar()
    fecha_nacimiento_var = tk.StringVar()
    edad_var = tk.StringVar()
    sexo_var = tk.StringVar()
    examen_var = tk.StringVar()
    resultado_var = tk.StringVar()

    # Registro de Pacientes
    formulario_paciente = ttk.LabelFrame(registro_pacientes, text="Datos del Paciente")
    formulario_paciente.grid(column=0, row=0, padx=10, pady=10, sticky='w')
    ttk.Label(formulario_paciente, text="Nombre completo:").grid(column=0, row=0, padx=10, pady=5, sticky='w')
    nombre_entry = ttk.Entry(formulario_paciente, textvariable=nombre_var, width=50)
    nombre_entry.grid(column=1, row=0, padx=10, pady=5)
    ttk.Label(formulario_paciente, text="Rut:").grid(column=0, row=1, padx=10, pady=5, sticky='w')
    rut_entry = ttk.Entry(formulario_paciente, textvariable=rut_var, width=50)
    rut_entry.grid(column=1, row=1, padx=10, pady=5)
    ttk.Label(formulario_paciente, text="Fecha de nacimiento (dd/mm/aaaa):").grid(column=0, row=2, padx=10, pady=5, sticky='w')
    fecha_nacimiento_entry = ttk.Entry(formulario_paciente, textvariable=fecha_nacimiento_var, width=50)
    fecha_nacimiento_entry.grid(column=1, row=2, padx=10, pady=5)
    ttk.Label(formulario_paciente, text="Edad:").grid(column=0, row=3, padx=10, pady=5, sticky='w')
    edad_entry = ttk.Entry(formulario_paciente, textvariable=edad_var, width=50)
    edad_entry.grid(column=1, row=3, padx=10, pady=5)
    ttk.Label(formulario_paciente, text="Sexo (F/M):").grid(column=0, row=4, padx=10, pady=5, sticky='w')
    sexo_combobox = ttk.Combobox(formulario_paciente, textvariable=sexo_var, values=["F", "M"], width=47)
    sexo_combobox.grid(column=1, row=4, padx=10, pady=5)

    guardar_button = ttk.Button(formulario_paciente, text="Guardar Paciente", command=lambda: guardar_paciente(root))
    guardar_button.grid(column=1, row=5, padx=10, pady=5)
    eliminar_button = ttk.Button(formulario_paciente, text="Eliminar Paciente", command=eliminar_paciente)
    eliminar_button.grid(column=2, row=3, padx=10, pady=5)
    editar_button = ttk.Button(formulario_paciente, text="Editar Paciente", command=editar_paciente)
    editar_button.grid(column=2, row=1, padx=10, pady=5)
    cargar_datos_button = ttk.Button(formulario_paciente, text="Cargar Datos del Paciente", command=cargar_datos_paciente_seleccionado)
    cargar_datos_button.grid(column=2, row=5, padx=10, pady=5)

    ttk.Label(registro_pacientes, text="Lista de Pacientes").grid(column=0, row=1, padx=10, pady=5, sticky='w')
    global lista_pacientes
    lista_pacientes = tk.Listbox(registro_pacientes, width=80, height=10)
    lista_pacientes.grid(column=0, row=2, padx=10, pady=5)

    # Registro de Exámenes
    formulario_examen = ttk.LabelFrame(registro_examenes, text="Agregar Exámenes")
    formulario_examen.grid(column=0, row=0, padx=10, pady=10, sticky='w')
    ttk.Label(formulario_examen, text="Seleccionar Paciente:").grid(column=0, row=0, padx=10, pady=5, sticky='w')
    global lista_pacientes_examen
    lista_pacientes_examen = ttk.Combobox(formulario_examen, values=lista_pacientes.get(0, tk.END), width=60)
    lista_pacientes_examen.grid(column=1, row=0, padx=10, pady=5)
    ttk.Label(formulario_examen, text="Buscar Examen:").grid(column=0, row=1, padx=10, pady=5, sticky='w')
    buscar_examen_var = tk.StringVar()
    buscar_examen_entry = ttk.Entry(formulario_examen, textvariable=buscar_examen_var, width=60)
    buscar_examen_entry.grid(column=1, row=1, padx=10, pady=5)
    buscar_examen_entry.bind("<KeyRelease>", lambda event: filtrar_examenes(event))
    ttk.Label(formulario_examen, text="Examen a agregar:").grid(column=0, row=2, padx=10, pady=5, sticky='w')
    global examenes_combobox
    examenes_combobox = ttk.Combobox(formulario_examen, textvariable=examen_var, width=60)
    examenes_combobox.grid(column=1, row=2, padx=10, pady=5)
    ttk.Label(formulario_examen, text="Resultado:").grid(column=0, row=3, padx=10, pady=5, sticky='w')
    resultado_entry = ttk.Entry(formulario_examen, textvariable=resultado_var, width=60)
    resultado_entry.grid(column=1, row=3, padx=10, pady=5)
    agregar_examen_button = ttk.Button(formulario_examen, text="Agregar Examen al Paciente", command=agregar_examen_paciente)
    agregar_examen_button.grid(column=1, row=4, padx=10, pady=5)

    ttk.Label(registro_examenes, text="Lista de Exámenes").grid(column=0, row=1, padx=10, pady=5, sticky='w')
    global lista_examenes
    lista_examenes = tk.Listbox(registro_examenes, width=70, height=10)
    lista_examenes.grid(column=0, row=2, padx=10, pady=5)
    eliminar_examen_button = ttk.Button(registro_examenes, text="Eliminar Examen", command=eliminar_examen)
    eliminar_examen_button.grid(column=0, row=5, padx=10, pady=5, sticky='w')

    # Listado de Derivación
    generar_listado_button = ttk.Button(listado_derivacion, text="Generar Listado de Derivación", command=generar_listado_con_pacientes)
    generar_listado_button.pack(padx=10, pady=10)

    # Botones adicionales: Programar cita y ver reporte
    ttk.Button(registro_pacientes, text="Programar Cita", command=programar_cita).grid(column=2, row=7, padx=10, pady=5)
    ttk.Button(registro_pacientes, text="Ver Reporte", command=reports.generate_statistics).grid(column=2, row=8, padx=10, pady=5)

    # Acerca de
    acerca_de_text = tk.Label(acerca_de, text="Sistema (R.P.E.D)\n\nVersión 2.0\n\nDesarrollado por //Gian Lucas San Martin//", font=("Arial", 19))
    acerca_de_text.pack(padx=50, pady=50)

    style = ttk.Style()
    style.configure('TButton', font=('Inconsolata', 14))

# Función para eliminar un paciente
def eliminar_paciente():
    try:
        index = lista_pacientes.curselection()[0]
        paciente_seleccionado = lista_pacientes.get(index)
        codigo_paciente = int(paciente_seleccionado.split(" - ")[0])
        del pacientes[codigo_paciente]
        lista_pacientes.delete(index)
        actualizar_lista_pacientes_examen()
    except IndexError:
        messagebox.showerror("Error", "Por favor, seleccione un paciente.")

# Función para editar un paciente
def editar_paciente():
    try:
        index = lista_pacientes.curselection()[0]
        paciente_seleccionado = lista_pacientes.get(index)
        codigo_paciente = int(paciente_seleccionado.split(" - ")[0])
        paciente = pacientes[codigo_paciente]
        
        # Actualizar los datos del paciente con los valores en los campos
        paciente['Nombre'] = nombre_var.get()
        paciente['Rut'] = rut_var.get()
        paciente['Fecha Nacimiento'] = fecha_nacimiento_var.get()
        paciente['Edad'] = edad_var.get()
        paciente['Sexo'] = sexo_var.get()
        
        actualizar_lista_pacientes()
        messagebox.showinfo("Éxito", "Datos del paciente actualizados.")
    except IndexError:
        messagebox.showerror("Error", "Por favor, seleccione un paciente.")

# Función para cargar los datos de un paciente seleccionado
def cargar_datos_paciente_seleccionado():
    try:
        paciente_seleccionado = lista_pacientes.get(lista_pacientes.curselection()[0])
        codigo_paciente = int(paciente_seleccionado.split(" - ")[0])
        paciente = pacientes[codigo_paciente]
        
        nombre_var.set(paciente['Nombre'])
        rut_var.set(paciente['Rut'])
        fecha_nacimiento_var.set(paciente['Fecha Nacimiento'])
        edad_var.set(paciente['Edad'])
        sexo_var.set(paciente['Sexo'])
    except IndexError:
        messagebox.showerror("Error", "Por favor, seleccione un paciente.")

# Función para filtrar los exámenes mientras se escribe en el campo de búsqueda
def filtrar_examenes(event):
    busqueda = event.widget.get().strip().lower()[:4]
    examenes_filtrados = [ex for ex in examenes_disponibles if busqueda in ex[1].strip().lower()]
    examenes_combobox['values'] = examenes_filtrados

# Función para agregar un examen a un paciente
def agregar_examen_paciente():
    try:
        paciente_seleccionado = lista_pacientes_examen.get()
        examen = examen_var.get()
        resultado = resultado_var.get()
        
        if paciente_seleccionado and examen:
            codigo_paciente = int(paciente_seleccionado.split(" - ")[0])
            codigo_barras = obtener_codigo_barras()
            pacientes[codigo_paciente]['Examenes'].append({
                'Examen': examen,
                'Código de Barras': codigo_barras,
                'Resultado': resultado
            })
            examen_var.set("")
            resultado_var.set("")
            actualizar_lista_examenes(codigo_paciente)
        else:
            messagebox.showerror("Error", "Seleccione un paciente y un examen válidos.")
    except Exception as e:
        messagebox.showerror("Error", f"Error al agregar el examen: {e}")

# Función para eliminar un examen de un paciente
def eliminar_examen():
    try:
        index = lista_examenes.curselection()[0]
        paciente_seleccionado = lista_pacientes_examen.get()
        codigo_paciente = int(paciente_seleccionado.split(" - ")[0])
        del pacientes[codigo_paciente]['Examenes'][index]
        lista_examenes.delete(index)
    except IndexError:
        messagebox.showerror("Error", "Seleccione un examen para eliminar.")

# Función para generar un listado de pacientes
def generar_listado_con_pacientes():
    carpeta = filedialog.askdirectory(title="Selecciona la carpeta de destino")
    if carpeta:
        generar_listado(pacientes, carpeta)
    else:
        messagebox.showinfo("Información", "No se seleccionó ninguna carpeta. Operación cancelada.")

# Función para programar una cita
def programar_cita():
    from tkinter import simpledialog
    codigo = simpledialog.askinteger("Programar Cita", "Ingrese el código del paciente:")
    if codigo is None or codigo not in pacientes:
        messagebox.showerror("Error", "Código de paciente inválido o no existente.")
        return
    nombre = pacientes[codigo]['Nombre']
    fecha = simpledialog.askstring("Programar Cita", "Ingrese la fecha y hora (dd/mm/aaaa hh:mm):")
    if not fecha:
        messagebox.showerror("Error", "Debe ingresar la fecha y hora.")
        return
    notes = simpledialog.askstring("Programar Cita", "Notas adicionales (opcional):")
    appointments.add_appointment(codigo, nombre, fecha, notes)
    messagebox.showinfo("Éxito", "Cita programada exitosamente.")

# Función para actualizar la lista de pacientes
def actualizar_lista_pacientes():
    lista_pacientes.delete(0, tk.END)
    for codigo, datos in pacientes.items():
        lista_pacientes.insert(tk.END, f"{codigo} - {datos['Nombre']}")
    actualizar_lista_pacientes_examen()

# Función para actualizar la lista de exámenes de un paciente
def actualizar_lista_pacientes_examen():
    global lista_pacientes_examen
    lista = [f"{codigo} - {datos['Nombre']}" for codigo, datos in pacientes.items()]
    lista_pacientes_examen['values'] = lista

def actualizar_lista_examenes(codigo_paciente):
    lista_examenes.delete(0, tk.END)
    for examen in pacientes[codigo_paciente]['Examenes']:
        lista_examenes.insert(tk.END, f"{examen['Examen']} - Código de Barras: {examen['Código de Barras']} - Resultado: {examen['Resultado']}")

def generar_listado(pacientes_data, carpeta_destino):
    try:
        fecha_actual = datetime.now().strftime('%d-%m-%Y_%H-%M-%S')
        nombre_archivo = f"listado_derivacion_{fecha_actual}.xlsx"
        archivo_excel = os.path.join(carpeta_destino, nombre_archivo)
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Listado de Derivación"
        
        sheet.cell(row=2, column=1, value="Fecha Actual")
        sheet.cell(row=2, column=2, value=datetime.now().strftime('%d/%m/%Y'))
        sheet.cell(row=4, column=1, value="Código de Cliente")
        sheet.cell(row=4, column=2, value="474")
        
        headers = ["Código Paciente", "Nombre Paciente", "Edad", "Examen", "Rut", "Sexo",
                   "F.U.R.", "Fecha Nac.", "F.U.D.", "Hora FUD", "Fecha T.M.", "Hora T.M.", "VOL"]
        sheet = workbook.active
        sheet.title = "Listado de Derivación"

        # Insertar encabezados
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=6, column=col, value=header)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Insertar datos
        row = 7
        for codigo, datos in pacientes_data.items():
            for examen in datos['Examenes']:
                sheet.cell(row=row, column=1, value=codigo)
                sheet.cell(row=row, column=2, value=datos['Nombre'].upper())
                sheet.cell(row=row, column=3, value=str(datos['Edad']).upper())
                sheet.cell(row=row, column=4, value=examen['Examen'].upper())
                sheet.cell(row=row, column=5, value=datos['Rut'].upper())
                sheet.cell(row=row, column=5, value=datos['Rut'].upper())
                sheet.cell(row=row, column=6, value=datos['Sexo'].upper())
                sheet.cell(row=row, column=7, value="")
                sheet.cell(row=row, column=8, value=datos['Fecha Nacimiento'])
                sheet.cell(row=row, column=9, value="")
                sheet.cell(row=row, column=10, value="")
                sheet.cell(row=row, column=11, value=datetime.now().strftime('%d/%m/%Y'))
                sheet.cell(row=row, column=12, value=datetime.now().strftime('%H:%M'))
                sheet.cell(row=row, column=13, value=examen['Resultado'])
                row += 1
        
        workbook.save(os.path.join(carpeta_destino, nombre_archivo))
        messagebox.showinfo("Éxito", f"Listado generado con éxito en: {nombre_archivo}")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo generar el listado:\n{e}")

# ---------------------- FUNCIONES DE GESTIÓN ----------------------
def guardar_paciente(parent):
    global pacientes
    nombre = nombre_var.get()
    rut = rut_var.get()
    fecha_nacimiento = fecha_nacimiento_var.get()
    edad = edad_var.get()
    sexo = sexo_var.get()

    if not re.match(r'^[A-Za-zÁÉÍÓÚáéíóúñÑ\s]+$', nombre):
        messagebox.showerror("Error", "El nombre solo puede contener letras y espacios")
        return
    if not fecha_nacimiento or not re.match(r'\d{2}/\d{2}/\d{4}', fecha_nacimiento):
        messagebox.showerror("Error", "Formato de fecha de nacimiento no válido. Debe ser 'dd/mm/aaaa'.")
        return
    if not sexo or not sexo.isalpha():
        messagebox.showerror("Error", "El sexo debe contener solo letras.")
        return
    if not rut or not validar_rut(rut):
        messagebox.showerror("Error", "Formato de RUT no válido.")
        return

    codigo_paciente = obtener_codigo_paciente(parent)
    if codigo_paciente is None:
        messagebox.showerror("Error", "Debe ingresar un número de paciente válido.")
        return

    paciente = {
        'Código': codigo_paciente,
        'Nombre': nombre,
        'Rut': rut,
        'Fecha Nacimiento': fecha_nacimiento,
        'Edad': edad,
        'Sexo': sexo,
        'Examenes': []
    }
    pacientes[codigo_paciente] = paciente
    lista_pacientes.insert(tk.END, f"{codigo_paciente} - {nombre}")
    actualizar_lista_pacientes_examen()

    nombre_var.set("")
    rut_var.set("")
    fecha_nacimiento_var.set("")
    edad_var.set("")
    sexo_var.set("")

    messagebox.showinfo("Éxito", "El paciente ha sido guardado en la base de datos.")
